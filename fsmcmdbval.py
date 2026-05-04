import sys,re, base64, urllib.request, urllib.parse, urllib.error, ssl, time, httplib2
from xml.dom.minidom import Node, parseString, parse
import xml.etree.ElementTree as ET
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import argparse

#----Funciones Generales
def read_list(input):
    txt = open(input, "r")
    resultado = []
    for element in txt:
        if "\\n" in element:
            element = element.replace("\\n", "")
        resultado.append(element)
    return resultado

def save_results(data):
    """ Guardar resultados finales """
    print (data)
    file = open("resultado.csv", "w")
    file.write(data)

def parse_xml(xml_data):

    cmdb = {}

    # limit = 100

    try:
        # Parsear el XML
        root = ET.fromstring(xml_data)


        for device in root.findall(".//device"):
            # limit -= 1
            # if limit == 0:
            #     return cmdb
            #print (device.text)
            device_ip = device.find("accessIp").text
            device_name = device.find("name").text
            flag_approve = device.find("approved").text
            vendor = device.find("deviceType//vendor").text  
            model = device.find("deviceType//model").text  

            device_type = f"{vendor} ({model})"
            cmdb[device_ip] = [device_name,flag_approve,device_type]

        return cmdb
    except ET.ParseError as e:
        print(f"Error al parsear el XML: {e}")

def clearwindow():
    ''' Esta funcion limpia la ventana'''
    from os import name, system
    if name == 'nt':
        system("cls")
    else:
        system("clear")



#----Funciones de extraccion de CMDB
def getCMDBInfo(appServer, user, password):
    base64_bytes = base64.encodebytes((user + ':' + password).encode())
    auth = 'Basic %s' % base64_bytes.decode()

    url = 'https://' + appServer + '/phoenix/rest/cmdbDeviceInfo/devices'
    request = urllib.request.Request(url)
    request.add_header('Authorization', auth.rstrip())
    request.add_header('User-Agent', 'Python-urllib2/2.7')
    request.get_method = lambda: 'GET'
    try:
        ssl._create_default_https_context = ssl._create_unverified_context
        ctx = ssl.SSLContext(ssl.PROTOCOL_TLSv1_2)
        clearwindow()
        opener = urllib.request.build_opener(urllib.request.HTTPSHandler(debuglevel=False, context=ctx))
        urllib.request.install_opener(opener)
        handle = urllib.request.urlopen(request)
        outXML = handle.read()
        return (outXML.decode())
    except urllib.error.HTTPError as e:
        if (e.code != 204):
            print (e)

#Funciones de Excel

# ─── Colores ───────────────────────────────────
C_ORANGE = "FF5A00"
C_DARK   = "1A1F36"
C_PANEL  = "252B4A"
C_WHITE  = "FFFFFF"
C_GREEN  = "2E7D32"
C_RED    = "C62828"
C_ROW_A  = "F7F9FC"
C_ROW_B  = "FFFFFF"


# ─── Utilidades de estilo ──────────────────────

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=10, color=C_DARK, name="Arial", italic=False):
    return Font(name=name, bold=bold, size=size, color=color, italic=italic)

def _align(h="left", v="center", wrap=False, indent=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)

def _thin_border():
    s = Side(style="thin", color="E0E4EF")
    return Border(left=s, right=s, top=s, bottom=s)


# ─── Normalización de eventos ──────────────────

def _normalize(ip, val):
    """
    Acepta el formato plano y el anidado que produce tu script:
      Plano:   { 'events': bool, 'event_types': ..., 'event_protocol': ... }
      Anidado: { 'IP':     { 'events': bool, ... } }
    """
    if isinstance(val, dict):
        if ip in val and isinstance(val[ip], dict):
            val = val[ip]
        if "events" in val:
            return val
    return {"events": False, "event_types": "Nologs", "event_protocol": "Nologs"}

def _fmt(value):
    if isinstance(value, list):
        return ", ".join(value) if value else "Sin tipo"
    return str(value)


# ─── Construcción de hojas ─────────────────────

def _cover(wb, time_range, rows):
    ws = wb.active
    ws.title = "Resumen"
    ws.sheet_view.showGridLines = False

    for row in ws.iter_rows(min_row=1, max_row=50, min_col=1, max_col=12):
        for cell in row:
            cell.fill = _fill(C_DARK)

    for col in range(1, 13):
        ws.cell(1, col).fill = _fill(C_ORANGE)
    ws.row_dimensions[1].height = 6

    ws.merge_cells("B4:K4")
    ws["B4"].value     = "FortiSIEM CMDB Validator"
    ws["B4"].font      = _font(bold=True, size=28, color=C_WHITE)
    ws["B4"].alignment = _align(indent=1)
    ws.row_dimensions[4].height = 42

    ws.merge_cells("B5:K5")
    ws["B5"].value     = "Reporte de Validación de Integraciones CMDB"
    ws["B5"].font      = _font(size=14, color=C_ORANGE)
    ws["B5"].alignment = _align(indent=1)
    ws.row_dimensions[5].height = 26

    ws.merge_cells("B7:K7")
    ws["B7"].fill = _fill(C_ORANGE)
    ws.row_dimensions[7].height = 3

    total   = len(rows)
    sending = sum(1 for r in rows if r[4] == "Enviando Logs")

    info = [
        ("📅  Rango de Tiempo Analizado", time_range),
        ("🗓  Fecha de Generación",        datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("🖥  Total de Dispositivos",       str(total)),
        ("✅  Enviando Logs",               str(sending)),
        ("❌  Sin Logs",                    str(total - sending)),
    ]

    for i, (label, value) in enumerate(info, start=9):
        ws.merge_cells(f"B{i}:E{i}")
        ws.merge_cells(f"F{i}:K{i}")
        lc = ws.cell(i, 2, label)
        vc = ws.cell(i, 6, value)
        lc.font      = _font(bold=True, size=11, color=C_ORANGE)
        vc.font      = _font(size=11, color=C_WHITE)
        lc.fill      = _fill(C_PANEL)
        vc.fill      = _fill("1E2340")
        lc.alignment = _align(indent=1)
        vc.alignment = _align(indent=1)
        ws.row_dimensions[i].height = 22

    ws.merge_cells("B21:K21")
    ws["B21"].value     = ("ℹ  Generado automáticamente por FortiSIEM CMDB Validator. "
                           "Ver hoja 'Detalle' para el listado completo con filtros.")
    ws["B21"].font      = _font(italic=True, size=9, color="8899AA")
    ws["B21"].alignment = _align(indent=1)



    for col, w in zip("BCDEFGHIJK", [2, 30, 2, 2, 2, 40, 2, 2, 2, 2]):
        ws.column_dimensions[col].width = w


def _detail(wb, time_range, rows):
    ws = wb.create_sheet("Detalle")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:G1")
    ws["A1"].value     = f"FortiSIEM CMDB Validator  |  {time_range}"
    ws["A1"].font      = _font(bold=True, size=12, color=C_WHITE)
    ws["A1"].fill      = _fill(C_ORANGE)
    ws["A1"].alignment = _align(indent=1)
    ws.row_dimensions[1].height = 24

    headers    = ["Dirección IP", "Hostname (CMDB)", "Aprobado", "Tipo de Dispositivo",
                  "Estado de Logs", "Tipos de Eventos", "Protocolo de Integración"]
    col_widths = [18, 34, 12, 24, 18, 30, 28]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell           = ws.cell(2, col_idx, h)
        cell.font      = _font(bold=True, size=10, color=C_WHITE)
        cell.fill      = _fill(C_DARK)
        cell.alignment = _align("center", wrap=True)
        cell.border    = Border(
            bottom=Side(style="medium", color=C_ORANGE),
            left=Side(style="thin", color="2D3561"),
            right=Side(style="thin", color="2D3561"),
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[2].height = 30

    border = _thin_border()
    for r_idx, row in enumerate(rows, start=3):
        is_sending = row[4] == "Enviando Logs"
        row_fill   = _fill("E8F5E9") if is_sending else (_fill(C_ROW_A) if r_idx % 2 == 0 else _fill(C_ROW_B))

        for c_idx, val in enumerate(row, start=1):
            cell           = ws.cell(r_idx, c_idx, val)
            cell.fill      = row_fill
            cell.border    = border
            cell.font      = _font()
            cell.alignment = _align(indent=1, wrap=True)

            if c_idx == 1:
                cell.font = Font(name="Courier New", size=9, color=C_DARK)
            elif c_idx == 5:
                cell.font      = _font(bold=True, color=C_GREEN if is_sending else C_RED)
                cell.alignment = _align("center")
            elif c_idx == 3:
                cell.alignment = _align("center")

        ws.row_dimensions[r_idx].height = 18
   
    for col_idx, col_cells in enumerate(ws.iter_cols(min_row=2, max_row=len(rows) +2), start=1):
            max_len = max((len(str(cell.value or "")) for cell in col_cells), default=10)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)

    table = Table(displayName="CMDBDetalle", ref=f"A2:G{len(rows) + 2}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showRowStripes=True, showColumnStripes=False,
        showFirstColumn=False, showLastColumn=False,
    )
    ws.add_table(table)


def _stats(wb, time_range, rows):
    ws = wb.create_sheet("Estadísticas")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("B2:H2")
    ws["B2"].value     = "Estadísticas del Reporte"
    ws["B2"].font      = _font(bold=True, size=18)
    ws["B2"].alignment = _align()
    ws.row_dimensions[2].height = 36

    ws.merge_cells("B3:H3")
    ws["B3"].value     = f"Período: {time_range}"
    ws["B3"].font      = _font(italic=True, size=10, color="666666")
    ws.row_dimensions[3].height = 18

    total   = len(rows)
    sending = sum(1 for r in rows if r[4] == "Enviando Logs")
    nologs  = total - sending
    pct     = f"{round(sending / total * 100, 1)}%" if total else "0%"

    kpis = [
        ("Total Dispositivos", total,   C_DARK,   C_WHITE),
        ("Enviando Logs ✅",   sending, "2E7D32", C_WHITE),
        ("Sin Logs ❌",        nologs,  "C62828", C_WHITE),
        ("Cobertura",          pct,     "E65100", C_WHITE),
    ]

    for i, (label, value, bg, fg) in enumerate(kpis):
        col = 2 + i * 2
        for row_n, val, sz in [(5, label, 10), (6, value, 16)]:
            ws.merge_cells(start_row=row_n, start_column=col, end_row=row_n, end_column=col + 1)
            cell           = ws.cell(row_n, col, val)
            cell.fill      = _fill(bg)
            cell.font      = _font(bold=True, size=sz, color=fg)
            cell.alignment = _align("center")
    ws.row_dimensions[5].height = 20
    ws.row_dimensions[6].height = 40

    protocols = {}
    for r in rows:
        for p in r[6].split(", "):
            p = p.strip()
            protocols[p] = protocols.get(p, 0) + 1

    ws.row_dimensions[9].height = 22
    for col_n, label in [(2, "Protocolo"), (3, "Dispositivos")]:
        cell           = ws.cell(9, col_n, label)
        cell.font      = _font(bold=True, size=11, color=C_WHITE)
        cell.fill      = _fill(C_ORANGE)
        cell.alignment = _align("center")

    border = _thin_border()
    for ri, (proto, count) in enumerate(sorted(protocols.items(), key=lambda x: -x[1]), start=10):
        row_fill = _fill(C_ROW_A) if ri % 2 == 0 else _fill(C_ROW_B)
        for col_n, val in [(2, proto), (3, count)]:
            cell           = ws.cell(ri, col_n, val)
            cell.fill      = row_fill
            cell.font      = _font()
            cell.alignment = _align("center")
            cell.border    = border
        ws.row_dimensions[ri].height = 18

    for col, w in zip("BCDEFGHI", [28, 16, 28, 16, 28, 16, 16, 16]):
        ws.column_dimensions[col].width = w


# ─── Función principal ─────────────────────────

def generate_report(events_data: dict, cmdb_data: dict, time_range: str,
                    output_path: str = "fortisiem_cmdb_report.xlsx"):
    """
    Genera el reporte XLSX del FortiSIEM CMDB Validator.

    Parámetros
    ----------
    events_data : dict
        Resultado del análisis de eventos por IP. Estructura esperada:
        { 'IP': {'events': bool, 'event_types': str|list, 'event_protocol': str|list} }
        También acepta el formato anidado que produce tu script:
        { 'IP': { 'IP': {'events': bool, 'event_types': list, 'event_protocol': list} } }

    cmdb_data : dict
        Datos del CMDB enriquecido. Estructura esperada:
        { 'IP': [hostname, approved_flag, device_type] }

    time_range : str
        Descripción del rango de tiempo analizado.
        Ejemplo: "Últimas 24 horas (2025-07-10 00:00 – 2025-07-10 23:59)"

    output_path : str
        Ruta donde se guarda el .xlsx. Por defecto: "fortisiem_cmdb_report.xlsx"
    """

    events_flat = {}
    for d in events_data:
        events_flat.update(d)

    all_ips = sorted(set(cmdb_data.keys()))
    rows = []
    for ip in all_ips:
        ev          = _normalize(ip, events_flat.get(ip, {}))
        cmdb        = cmdb_data.get(ip, [None, None, None])
        hostname    = cmdb[0] if cmdb[0] is not None else "N/A"
        approved    = str(cmdb[1]) if cmdb[1] is not None else "N/A"
        device_type = str(cmdb[2]) if cmdb[2] is not None else "N/A"
        is_sending  = ev.get("events", False)
        status      = "Enviando Logs" if is_sending else "Sin Logs"
        event_types = _fmt(ev.get("event_types", "Nologs"))
        protocol    = _fmt(ev.get("event_protocol", "Nologs"))
        rows.append([ip, hostname, approved, device_type, status, event_types, protocol])

    wb = Workbook()
    _cover(wb, time_range, rows)
    _detail(wb, time_range, rows)
    _stats(wb, time_range, rows)

    wb.save(output_path)
    print(f"[✓] Reporte generado: {output_path}  ({len(rows)} dispositivos)")

#----FUnciones de extraccion de eventos
def extrat_data_status(xml_string):
    query = ET.fromstring(xml_string).find('./result/progress').text
    return int(query)

def extrat_data_query(xml):
    query = (ET.fromstring(xml).get('requestId'))
    timestamp = ET.fromstring(xml).find('./result/expireTime').text
    resultado = f"{query},{timestamp}"
    return resultado


def dumpXML(xmlList):
    param = []
    for xml in xmlList:
        doc = parseString(xml.encode('ascii', 'xmlcharrefreplace'))
    for node in doc.getElementsByTagName("events"):
        for node1 in node.getElementsByTagName("event"):
            mapping = {}
            for node2 in node1.getElementsByTagName("attributes"):
                for node3 in node2.getElementsByTagName("attribute"):
                    itemName = node3.getAttribute("name")
                    for node4 in node3.childNodes:
                        if node4.nodeType == Node.TEXT_NODE:
                            message = node4.data
                            if '\\n' in message:
                                message = message.replace('\\n', '')
                            mapping[itemName] = message
            param.append(mapping)
    return param

def select_query(input_time, ip_device):

    timestamp_now = int(time.time())

    custom_time = 60 * input_time # 1 hora
    
    time_start = timestamp_now - custom_time
    time_end = timestamp_now - custom_time + 3600

    last_event = f"""<?xml version="1.0" encoding="UTF-8"?>
    <Reports>
        <Report baseline="" rsSync="">
            <Name>FortiSIEM CMDB Validator</Name>
            <CustomerScope>
                <Include all="true">
                </Include>
                <Exclude>
                </Exclude>
            </CustomerScope>
            <SelectClause>
                <AttrList>extEventRecvProto,eventType,COUNT(phRecvTime)</AttrList>
            </SelectClause>
            <OrderByClause>
                <AttrList>COUNT(*) DESC</AttrList>
            </OrderByClause>
            <PatternClause window="3600">
                <SubPattern name="Filter">
                    <SingleEvtConstr>(reptDevIpAddr = {ip_device} AND eventType NOT CONTAIN "PH_")</SingleEvtConstr>
                    <GroupByAttr>eventType,extEventRecvProto</GroupByAttr>
                </SubPattern>
            </PatternClause>
            <ReportInterval>
                <Low>{time_start}</Low>
                <High>{time_end}</High>
            </ReportInterval>
        </Report>
</Reports>"""
    return last_event


def get_queryfromsiem(ip_siem, user, password, input_time,ip_device):

   
    xml_query = select_query(input_time, ip_device)

    url = "https://" + ip_siem + ":443/phoenix/rest/query/"
    urlfirst = url + "eventQuery"
    
    h = httplib2.Http(disable_ssl_certificate_validation=True)
    h.add_credentials(user, password)
    
    header = {'Content-Type': 'text/xml'}
    
    doc = parseString(xml_query)
    t = doc.toxml()

    if '<DataRequest' in t:
        t1 = t.replace("<DataRequest", "<Reports><Report")
    else:
        t1 = t
    if '</DataRequest>' in t1:
        t2 = t1.replace("</DataRequest>", "</Report></Reports>")
    else:
        t2 = t1

    resp, content = h.request(urlfirst, "POST", t2, header)
    queryId = content.decode("utf-8")
    # print (t2)
    # print (queryId)
    if "xml version" in queryId:
        queryId = extrat_data_query(queryId)
        
    if 'error code="255"' in queryId:
        print ("Query Error, check sending XML file.")
        exit()

    urlSecond = url + "progress/" + queryId
    if resp['status'] == '200':
        resp, content = h.request(urlSecond)

    else:
        print ("appServer doesn't return query. Error code is %s" % resp['status'] )
        exit()


    while True:
        resp, content = h.request(urlSecond)
        try: 
            progreso = extrat_data_status(content)
            #print (progreso)
            if progreso == 100:
                break
        except:
            while content.decode("utf-8") != '100':
                resp, content = h.request(urlSecond)
            break

    outXML = []
   
    
    urlFinal = url + 'events/' + queryId + '/0/1000'
    resp, content = h.request(urlFinal)
    
    #print (resp, "\\n\\n", content)

    if content != '':
        outXML.append(content.decode("utf-8"))

    p = re.compile('totalCount="\\d+"')
    mlist = p.findall(content.decode())
    
    
    if mlist[0] != '':
        mm = mlist[0].replace('"', '')
        m = mm.split("=")[-1]
        num = 0
        if int(m) > 1000:
            num = int(m) / 1000
            if int(m) % 1000 > 0:
                num += 1
        if num > 0:
            for i in range(num):
                urlFinal = url + 'events/' + queryId + '/' + str((i + 1) * 1000) + '/1000'
                resp, content = h.request(urlFinal)
                if content != '':
                    outXML.append(content.decode("utf-8"))
    else:
        print ("no info in this report.")
        return "Error"
    data = dumpXML(outXML)


    result= detect_eventtypes(data,ip_device)
    return result


EVENT_TYPE_PATTERNS = {

    # ── Windows ──────────────────────────────────────────
    "win-security"        : "Windows Security Events",
    "win-sysmon"          : "Windows Sysmon Events",
    "win-system"          : "Windows System Events",
    "win-application"     : "Windows Application Events Events",
    "win-powershell"      : "Windows PowerShell Events",
    "win-wmi"             : "Windows WMI Activity Events",
    "win-dns"             : "Windows DNS Server Events",
    "win-dhcp"            : "Windows DHCP Server Events",
    "win-taskscheduler"   : "Windows Task Scheduler Events",
    "win-bits"            : "Windows BITS Client Events",
    "win-firewall"        : "Windows Firewall Events",
    "win-defender"        : "Windows Defender Events",
    "win-applocker"       : "Windows AppLocker Events",
    "win-printservice"    : "Windows Print Service Events",
    "win-terminalservice" : "Windows Terminal Services Events",
    "win-remotedesktop"   : "Windows Remote Desktop Events",

    # ── Linux / Unix ──────────────────────────────────────
    "linux-auth"          : "Linux Authentication Events",
    "linux-syslog"        : "Linux Syslog Events",
    "linux-audit"         : "Linux Audit (auditd) Events",
    "linux-cron"          : "Linux Cron Jobs Events",
    "linux-kernel"        : "Linux Kernel Events",
    "linux-sudo"          : "Linux Sudo Activity Events",
    "linux-ssh"           : "Linux SSH Activity Events",
    "linux-rpm"           : "Linux Package Management Events",
    "linux-apt"           : "Linux Package Management Events",

    # ── Fortinet ──────────────────────────────────────────
    "fortigate-traffic"    : "Fortinet Firewall Traffic Events",
    "fortinet-utm"        :  "Fortinet UTM Events",
    "fortinet-vpn"        : "Fortinet VPN Events",
    "fortigate-ips"        : "Fortinet IPS/IDS Events",
    "fortigate-antivirus"  : "Fortinet Antivirus Events",
    "fortigate-webfilter"  : "Fortinet Web Filter Events",
    "fortigate-appctrl"    : "Fortinet Application Control Events",
    "fortigate-system"     : "Fortinet System Events",
    "fortigate-auth"       : "Fortinet Authentication Events",
    "fortigate-ha"         : "Fortinet High Availability Events",
    "faz"                 : "FortiAnalyzer Events",
    "fmg"                 : "FortiManager Events",
    "fortiedr"            : "FortiEDR Events",
    "forticlient"         : "FortiClient Endpoint Events",
    "fortimail"           : "FortiMail Events Events",
    "fortiweb"            : "FortiWeb WAF Events",
    "fortisandbox"        : "FortiSandbox Events",
    "fortiauth"           : "FortiAuthenticator Events",
    "fortiswitch"         : "FortiSwitch Events",
    "fortirecon-easm-"    : "FortiRecon ESASM Events",
    "fortiwifi"           : "FortiWiFi / FortiAP Events",
    "fortideceptor-scada-alert"  : "FortiDeceptor Scada Events Alerts",

    # ── Palo Alto ─────────────────────────────────────────
    "panos-traffic"       : "Palo Alto Firewall Traffic Events",
    "panos-threat"        : "Palo Alto Threat Events",
    "panos-system"        : "Palo Alto System Events",
    "panos-config"        : "Palo Alto Config Changes Events",
    "panos-auth"          : "Palo Alto Authentication Events",
    "panos-vpn"           : "Palo Alto GlobalProtect VPN Events",
    "panos-wildfire"      : "Palo Alto WildFire Sandbox Events",
    "panos-url"           : "Palo Alto URL Filtering Events",

    # ── Cisco ─────────────────────────────────────────────
    "cisco-asa"           : "Cisco ASA Firewall Events",
    "cisco-ftd"           : "Cisco Firepower (FTD) Events",
    "cisco-ios"           : "Cisco IOS Router/Switch Events",
    "cisco-nx"            : "Cisco NX-OS Events",
    "cisco-meraki"        : "Cisco Meraki Events",
    "cisco-ise"           : "Cisco ISE (NAC/Auth) Events",
    "cisco-umbrella"      : "Cisco Umbrella DNS Events",
    "cisco-amp"           : "Cisco AMP / Secure Endpoint Events",
    "cisco-vpn"           : "Cisco AnyConnect VPN Events",
    "cisco-wlc"           : "Cisco Wireless Controller Events",

    # ── Microsoft Cloud / Active Directory ───────────────
    "msad"                : "Microsoft Active Directory Events",
    "mso365"              : "Microsoft 365 Audit Events",
    "msazure"             : "Microsoft Azure Activity Events",
    "msentra"             : "Microsoft Entra ID (AAD) Events",
    "msexchange"          : "Microsoft Exchange Events",
    "msteams"             : "Microsoft Teams Events",
    "msdefender"          : "Microsoft Defender for Endpoint Events",
    "mssql"               : "Microsoft SQL Server Events",
    "mssharepoint"        : "Microsoft SharePoint Events",

    # ── AWS ──────────────────────────────────────────────
    "aws-cloudtrail"      : "AWS CloudTrail Events",
    "aws-guardduty"       : "AWS GuardDuty Events",
    "aws-securityhub"     : "AWS Security Hub Events",
    "aws-vpc"             : "AWS VPC Flow Logs Events",
    "aws-s3"              : "AWS S3 Access Logs Events",
    "aws-iam"             : "AWS IAM Events ",
    "aws-waf"             : "AWS WAF Events",

    # ── GCP ──────────────────────────────────────────────
    "gcp-audit"           : "GCP Audit Logs Events",
    "gcp-vpc"             : "GCP VPC Flow Logs Events",
    "gcp-iam"             : "GCP IAM Events",

    # ── Networking general ────────────────────────────────
    "netflow"             : "NetFlow / Traffic Flow Events",
    "sflow"               : "sFlow Traffic Events",
    "ipfix"               : "IPFIX Flow Events",
    "snmp"                : "SNMP Trap Events",
    "bgp"                 : "BGP Routing Events Events",
    "ospf"                : "OSPF Routing Events Events",
    "dhcp"                : "DHCP Events Events",
    "dns"                 : "DNS Query/Response",
    "ntp"                 : "NTP Events",

    # ── Seguridad / EDR / SIEM ────────────────────────────
    "crowdstrike"         : "CrowdStrike Falcon EDR",
    "sentinelone"         : "SentinelOne EDR",
    "carbonblack"         : "VMware Carbon Black EDR",
    "cybereason"          : "Cybereason EDR",
    "symantec-edr"        : "Symantec EDR",
    "trendmicro"          : "Trend Micro",
    "mcafee"              : "McAfee / Trellix",
    "tenable"             : "Tenable / Nessus Vulnerability",
    "qualys"              : "Qualys Vulnerability",
    "rapid7"              : "Rapid7 InsightVM",
    "darktrace"           : "Darktrace AI Detection",
    "vectra"              : "Vectra AI NDR",
    "nozomi"              : "Nozomi OT/ICS Security",
    "claroty"             : "Claroty OT Security",
    "armis"               : "Armis Asset Intelligence",
    "dragos"              : "Dragos OT Security",

    # ── Aplicaciones / Web ────────────────────────────────
    "apache"              : "Apache Web Server",
    "nginx"               : "Nginx Web Server",
    "iis"                 : "IIS Web Server",
    "tomcat"              : "Apache Tomcat",
    "f5"                  : "F5 BIG-IP",
    "imperva"             : "Imperva WAF",
    "akamai"              : "Akamai WAF/CDN",

    # ── Bases de datos ────────────────────────────────────
    "oracle-db"           : "Oracle Database",
    "mysql"               : "MySQL / MariaDB",
    "postgresql"          : "PostgreSQL",
    "mongodb"             : "MongoDB",
    "redis"               : "Redis",

    # ── Identidad / Acceso ────────────────────────────────
    "okta"                : "Okta Identity",
    "pingidentity"        : "Ping Identity",
    "beyondtrust"         : "BeyondTrust PAM",
    "cyberark"            : "CyberArk PAM",
    "sailpoint"           : "SailPoint IGA",
    "radius"              : "RADIUS Authentication",
    "ldap"                : "LDAP Directory",
    "kerberos"            : "Kerberos Authentication",

    # ── Correo ────────────────────────────────────────────
    "proofpoint"          : "Proofpoint Email Security",
    "mimecast"            : "Mimecast Email Security",
    "barracuda-email"     : "Barracuda Email Security",

    # ── OT / ICS ─────────────────────────────────────────
    "scada"               : "SCADA Events",
    "modbus"              : "Modbus Protocol",
    "dnp3"                : "DNP3 Protocol",
    "s7"                  : "Siemens S7 PLC",

}


# Ordenar de más largo a más corto para que patrones específicos ganen
PATTERNS = sorted(EVENT_TYPE_PATTERNS.keys(), key=len, reverse=True)

def get_category(event_type: str) -> str | None:
    """Retorna la categoría del event type o None si no hay match."""
    event_lower = event_type.lower()
    for pattern in PATTERNS:
        if event_lower.startswith(pattern):
            return EVENT_TYPE_PATTERNS[pattern]
    return None

def classify_device_events(event_types: list) -> list:
    """
    Recibe la lista de event_types de un dispositivo
    y retorna las categorías únicas identificadas.
    """
    identified = set()  # set en lugar de lista para evitar duplicados sin buscar
    for event_type in event_types:
        category = get_category(event_type)
        if category:
            identified.add(category)
    return list(identified)

def detect_eventtypes(data,ip_device):


    filter_detected = []
    protocol_detected = []

    if len(data) == 0:  #Sin eventos    
        return {"events":False, "event_types":"No logs","event_protocol": "No logs" }
    else:

        event_types = []
        for element in data:
            
            if element["eventType"] == "Unknown_EventType":
                filter_detected.append("Desconocido")
                protocol_detected.append("Desconocido")
            else:
                event_types.append(element["eventType"].lower())

                if element["extEventRecvProto"] not in protocol_detected:
                    protocol_detected.append(element["extEventRecvProto"])


        categories = classify_device_events(event_types)
        return {ip_device:{"events":True, "event_types":categories, "event_protocol":protocol_detected}}
        

def main():
    from concurrent.futures import ProcessPoolExecutor, as_completed
    
    #Comandos de configuracion
    parser = argparse.ArgumentParser(description="Options")
    parser.add_argument("-u", "--user", help="Set User account")
    parser.add_argument("-p", "--passw", help="Set Password account")
    parser.add_argument("-s","--siem", help="Set IP FortiSiEM Supervisor IP")
    parser.add_argument("-o", "--output", help="Set output file path")
    parser.add_argument("-t", "--time", help="Set time range logs")

    #Comandos de funciones
    parser.add_argument("-xall", "--extractall", action="store_true", help="Extraer la CMDB y toda la informacion relacionadas")
    parser.add_argument("-xnologs", "--extractnologs", action="store_true", help="Extraer solo los que no envian eventos")
 
    args = parser.parse_args()
    ip_siem = args.siem
    username = args.user
    password = args.passw

    if args.extractall:  #Extraer info de toda la CMDB

        resultados = []
        
        from tqdm import tqdm

        # Step 1: Extraer listado de equipos en la CMDB
        apiquery = getCMDBInfo(ip_siem, username, password)
        cmdb = parse_xml(apiquery)

        if None in cmdb:
            del cmdb[None]

        total_device = len(cmdb)

        print (f"TOTAL DE EQUIPOS: {total_device}")
        
        # Step 2: Validar envio de eventos al SIEM
        from functools import partial

        #get_queryfromsiem(ip_siem, username, password, select_query(input_time,ip_device))
        func = partial(get_queryfromsiem,ip_siem, username, password, int(args.time)
                        )

        with ProcessPoolExecutor(max_workers=8) as executor:

            device_analyzed = dict(
                zip (cmdb.keys(), tqdm( executor.map(func, cmdb.keys()),total=total_device))
                                )
            resultados.append(device_analyzed)
    
    # Generate Excel Report

    generate_report(events_data=resultados, cmdb_data=cmdb, time_range=f"Ultimos {args.time} minutos", output_path=args.output)

if __name__ == "__main__":
    main()
